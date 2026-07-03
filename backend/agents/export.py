import io
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXPECTED_HEADERS = [
    "Test Key", "Summary", "Type", "Component",
    "Description", "Action", "Data", "Expected Result", "Release",
]
COL_WIDTHS = {
    "Test Key": 12, "Summary": 96, "Type": 10, "Component": 17,
    "Description": 38, "Action": 130, "Data": 20, "Expected Result": 30, "Release": 9,
}


def _parse_steps(tc: dict) -> tuple[list[str], list[str], list[str]]:
    raw_action = str(tc.get("Action", "")).replace("\\n", "\n")
    steps = [s.strip() for s in re.split(r'\n|(?=\d+\.\s)', raw_action) if s.strip()] or [""]
    n = len(steps)

    raw_data = str(tc.get("Data", "")).strip()
    data_parts = [d.strip() for d in raw_data.split("|")] if "|" in raw_data else [raw_data]
    data_parts = [d for d in data_parts if d] or ["N/A"]
    while len(data_parts) < n:
        data_parts.append(data_parts[-1])

    raw_exp = str(tc.get("Expected Result", "")).strip()
    exp_parts = [e.strip() for e in raw_exp.split("|")] if "|" in raw_exp else [raw_exp]
    exp_parts = [e for e in exp_parts if e] or ["Result not specified"]
    while len(exp_parts) < n:
        exp_parts.append(exp_parts[-1])

    return steps, data_parts, exp_parts


def build_excel(test_cases: list) -> tuple[io.BytesIO, int]:
    if not test_cases:
        test_cases = [{h: "" for h in EXPECTED_HEADERS}]

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    lr_border = Border(left=thin, right=thin)

    header_font = Font(name="Aptos Narrow", size=11, bold=True)
    body_font = Font(name="Aptos Narrow", size=11)
    action_font = Font(name="Aptos Narrow", size=12)

    for col_idx, header in enumerate(EXPECTED_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
    ws.row_dimensions[1].height = 16

    current_row = 2
    for tc_idx, tc in enumerate(test_cases, 1):
        steps, data_parts, exp_parts = _parse_steps(tc)
        n = len(steps)

        for step_idx, step in enumerate(steps):
            row = current_row + step_idx
            ws.row_dimensions[row].height = 16

            if step_idx == 0:
                meta = {
                    "Test Key": tc_idx,
                    "Summary": str(tc.get("Summary", "")),
                    "Type": str(tc.get("Type", "Manual")),
                    "Component": str(tc.get("Component", "")),
                    "Description": str(tc.get("Description", "")),
                    "Release": str(tc.get("Release", "TBD")),
                }
                for col_idx, header in enumerate(EXPECTED_HEADERS, 1):
                    if header in meta:
                        cell = ws.cell(row=row, column=col_idx, value=meta[header])
                        cell.font = body_font
                        cell.border = thin_border
                        if header == "Summary":
                            cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                for col_idx in [1, 2, 3, 4, 5, 9]:
                    ws.cell(row=row, column=col_idx, value=None)

            ws.cell(row=row, column=6, value=step).font = action_font
            ws.cell(row=row, column=6).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=6).border = lr_border

            ws.cell(row=row, column=7, value=data_parts[step_idx]).font = body_font
            ws.cell(row=row, column=7).border = thin_border

            ws.cell(row=row, column=8, value=exp_parts[step_idx]).font = body_font
            ws.cell(row=row, column=8).border = thin_border

        current_row += n

    for col_idx, header in enumerate(EXPECTED_HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[header]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(test_cases)


def build_csv(test_cases: list) -> tuple[io.BytesIO, int]:
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL)
    writer.writerow(EXPECTED_HEADERS)

    for tc_idx, tc in enumerate(test_cases, 1):
        steps, data_parts, exp_parts = _parse_steps(tc)
        for step_idx, step in enumerate(steps):
            writer.writerow([
                tc_idx,
                str(tc.get("Summary", "")),
                str(tc.get("Type", "Manual")),
                str(tc.get("Component", "")),
                str(tc.get("Description", "")),
                step,
                data_parts[step_idx],
                exp_parts[step_idx],
                str(tc.get("Release", "TBD")),
            ])

    csv_bytes = buf.getvalue().encode("utf-8-sig")
    return io.BytesIO(csv_bytes), len(test_cases)
